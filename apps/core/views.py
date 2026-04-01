from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db import transaction

from apps.core.mixins import OwnerRequiredMixin, StaffRequiredMixin


def landing_view(request):
    """SEO-optimized public landing page."""
    if request.user.is_authenticated:
        if request.user.role == 'tenant':
            return redirect('tenant:home')
        return redirect('dashboard:index')
    return render(request, 'landing.html')


def login_view(request):
    """Login page — redirects based on role."""
    if request.user.is_authenticated:
        if request.user.role == 'tenant':
            return redirect('tenant:home')
        return redirect('dashboard:index')

    from django.contrib.auth.forms import AuthenticationForm
    form = AuthenticationForm(request)

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            next_url = request.POST.get('next') or request.GET.get('next')
            if next_url:
                return redirect(next_url)
            if user.role == 'tenant':
                return redirect('tenant:home')
            return redirect('dashboard:index')
        else:
            messages.error(request, _('Invalid username or password.'))

    return render(request, 'auth/login.html', {'form': form, 'next': request.GET.get('next', '')})


@require_POST
def logout_view(request):
    """Logout — POST only."""
    logout(request)
    messages.success(request, _('You have been logged out.'))
    return redirect('core:login')


class SetupWizardView(OwnerRequiredMixin, View):
    """3-step setup wizard: dormitory info → rooms → billing settings.
    เฉพาะ owner/superadmin เท่านั้นที่ตั้งค่าหอพักได้ — staff ไม่มีสิทธิ์
    """

    def _get_step(self, request):
        return int(request.GET.get('step', request.POST.get('step', 1)))

    def get(self, request, *args, **kwargs):
        step = self._get_step(request)
        return render(request, 'core/setup_wizard.html', {
            'step': step,
            'form': _WizardForm(step),
        })

    def post(self, request, *args, **kwargs):
        step = self._get_step(request)

        if step == 1:
            return self._handle_step1(request)
        elif step == 2:
            return self._handle_step2(request)
        elif step == 3:
            return self._handle_step3(request)

        return redirect('core:setup_wizard')

    def _handle_step1(self, request):
        from apps.core.models import Dormitory
        name = request.POST.get('name', '').strip()
        address = request.POST.get('address', '').strip()
        if not name or not address:
            messages.error(request, _('Please fill in all required fields.'))
            return render(request, 'core/setup_wizard.html', {
                'step': 1,
                'form': _WizardForm(1, request.POST),
            })

        dorm, created = Dormitory.objects.get_or_create(
            name=name,
            defaults={'address': address}
        )
        if not created:
            dorm.address = address
            dorm.save()

        if not request.user.dormitory:
            request.user.dormitory = dorm
            request.user.save()

        return redirect('core:setup_wizard' + '?step=2')

    def _handle_step2(self, request):
        from apps.rooms.models import Building, Floor, Room
        dorm = request.user.dormitory
        if not dorm:
            return redirect('core:setup_wizard')

        building_name = request.POST.get('building_name', '').strip()
        num_floors = int(request.POST.get('num_floors', 0) or 0)
        rooms_per_floor = int(request.POST.get('rooms_per_floor', 0) or 0)
        default_rent = request.POST.get('default_rent', 0) or 0

        if building_name and num_floors > 0 and rooms_per_floor > 0:
            building, _created = Building.objects.get_or_create(
                dormitory=dorm, name=building_name)
            for floor_num in range(1, num_floors + 1):
                floor, _created = Floor.objects.get_or_create(
                    building=building, number=floor_num)
                for room_num in range(1, rooms_per_floor + 1):
                    room_number = f"{floor_num}{str(room_num).zfill(2)}"
                    Room.objects.get_or_create(
                        floor=floor,
                        number=room_number,
                        defaults={'base_rent': default_rent}
                    )

        return redirect('core:setup_wizard' + '?step=3')

    def _handle_step3(self, request):
        from apps.billing.models import BillingSettings
        dorm = request.user.dormitory
        if not dorm:
            return redirect('core:setup_wizard')

        settings, _created = BillingSettings.objects.get_or_create(dormitory=dorm)
        settings.elec_rate = request.POST.get('elec_rate', 7.00)
        settings.water_rate = request.POST.get('water_rate', 18.00)
        settings.bill_day = int(request.POST.get('bill_day', 1))
        settings.grace_days = int(request.POST.get('grace_days', 5))
        settings.save()

        messages.success(request, _('Setup completed successfully!'))
        return redirect('dashboard:index')


@login_required
@require_POST
def theme_toggle_view(request):
    """Toggle user theme between light and dark."""
    from apps.core.models import CustomUser
    user = request.user
    user.theme = CustomUser.Theme.DARK if user.theme == CustomUser.Theme.LIGHT else CustomUser.Theme.LIGHT
    user.save(update_fields=['theme'])
    return redirect(request.POST.get('next') or request.META.get('HTTP_REFERER') or 'dashboard:index')


@login_required
@require_POST
def property_switch_view(request):
    """Switch the active dormitory context for staff/owner users.

    Validates that the requested dormitory is one the user belongs to before
    writing it to the session.  Redirects back to the referring page.
    """
    from apps.core.models import Dormitory, UserDormitoryRole
    from django.core.exceptions import ValidationError

    dormitory_id = request.POST.get('dormitory_id')
    if dormitory_id:
        try:
            dorm = Dormitory.objects.get(
                pk=dormitory_id,
                userdormitoryrole__user=request.user,
            )
            request.session['active_dormitory_id'] = str(dorm.pk)
        except (Dormitory.DoesNotExist, ValidationError):
            messages.error(request, _(
                'You do not have access to that property.'))

    return redirect(request.POST.get('next') or 'dashboard:index')


class AuditLogView(OwnerRequiredMixin, View):
    """
    แสดง Audit Log สำหรับ Owner — เห็นได้เฉพาะ log ของ tenant ตัวเอง
    Filter: date_from, date_to, user_id, model_name, action
    Pagination: 20 per page
    เฉพาะ owner/superadmin เท่านั้น — role อื่น return 403 (ผ่าน OwnerRequiredMixin)
    """

    def get(self, request):
        from apps.core.models import ActivityLog, CustomUser

        dormitory = request.active_dormitory

        # Base queryset — filter by dormitory (tenant isolation)
        qs = ActivityLog.unscoped_objects.select_related('user').order_by('-created_at')
        if dormitory:
            qs = qs.filter(dormitory=dormitory)
        elif request.user.role != 'superadmin':
            qs = qs.none()

        # Filters
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        user_filter = request.GET.get('user_id', '').strip()
        model_filter = request.GET.get('model_name', '').strip()
        action_filter = request.GET.get('action', '').strip()

        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        if user_filter:
            qs = qs.filter(user_id=user_filter)
        if model_filter:
            qs = qs.filter(model_name=model_filter)
        if action_filter:
            qs = qs.filter(action=action_filter)

        # Pagination
        paginator = Paginator(qs, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        # สร้าง dropdown choices สำหรับ filter
        # ดึง distinct model_names ที่มีใน log ของ dormitory นี้
        model_choices = (
            ActivityLog.unscoped_objects
            .filter(dormitory=dormitory)
            .exclude(model_name='')
            .values_list('model_name', flat=True)
            .distinct()
            .order_by('model_name')
            if dormitory else []
        )

        # ดึง users ที่มี log ใน dormitory นี้
        user_choices = (
            CustomUser.objects.filter(
                activitylog__dormitory=dormitory
            ).distinct().order_by('username')
            if dormitory else []
        )

        action_choices = [
            ('create', _('Create')),
            ('update', _('Update')),
            ('delete', _('Delete')),
        ]

        has_filter = any([date_from, date_to, user_filter, model_filter, action_filter])

        return render(request, 'core/audit_log.html', {
            'page_obj': page_obj,
            'date_from': date_from,
            'date_to': date_to,
            'user_filter': user_filter,
            'model_filter': model_filter,
            'action_filter': action_filter,
            'model_choices': list(model_choices),
            'user_choices': user_choices,
            'action_choices': action_choices,
            'has_filter': has_filter,
        })


# Backward-compatible alias สำหรับ URL conf ที่ reference audit_log_view โดยตรง
audit_log_view = AuditLogView.as_view()


# ---------------------------------------------------------------------------
# Data Import Wizard — Rooms
# ---------------------------------------------------------------------------

# คอลัมน์ที่ต้องมีใน Excel template สำหรับ import ห้อง
ROOM_IMPORT_COLUMNS = ['building_name', 'floor_number', 'room_number', 'room_type', 'base_rent', 'status']
# คอลัมน์ที่ต้องมีใน Excel template สำหรับ import ผู้เช่า
TENANT_IMPORT_COLUMNS = ['room_number', 'building_name', 'first_name', 'last_name', 'phone', 'email', 'line_id', 'start_date']

# สถานะห้องที่รับได้
VALID_ROOM_STATUSES = {'occupied', 'vacant', 'cleaning', 'maintenance'}


def _parse_room_excel(workbook, dormitory):
    """
    Parse ไฟล์ Excel สำหรับ import ห้องพัก
    คืนค่า (valid_rows: list[dict], error_rows: list[dict])
    - valid_rows: แถวที่ผ่าน validation ทั้งหมด พร้อม import
    - error_rows: แถวที่มี error พร้อม error message รายแถว
    ทั้งสองรายการ return พร้อมกันเสมอ (partial validation — ไม่ abort ทั้งหมดเมื่อมี error บางแถว)
    ยกเว้น: ถ้า header ไม่ครบ คืน ([], [{'row_num': 0, 'error': ...}]) ทันที
    """
    from apps.rooms.models import Building, Floor, Room

    ws = workbook.active
    headers = [cell.value for cell in ws[1]]

    # ตรวจสอบว่า header ครบถ้วน — ถ้าไม่ครบให้ abort ทันที ไม่อ่านข้อมูลต่อ
    missing_cols = [col for col in ROOM_IMPORT_COLUMNS if col not in headers]
    if missing_cols:
        fatal_error = _('Missing columns: %(cols)s') % {'cols': ', '.join(missing_cols)}
        return [], [{'row_num': 0, 'error': fatal_error}]

    col_idx = {col: headers.index(col) for col in ROOM_IMPORT_COLUMNS}
    valid_rows = []
    error_rows = []

    # ใช้ set ตรวจ duplicate ภายในไฟล์ก่อน (building, floor, room_number)
    seen_keys = set()

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # ข้าม row ว่างทั้งหมด
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        def get_val(col):
            idx = col_idx[col]
            return row[idx] if idx < len(row) else None

        building_name = str(get_val('building_name') or '').strip()
        floor_number = get_val('floor_number')
        room_number = str(get_val('room_number') or '').strip()
        room_type = str(get_val('room_type') or '').strip()
        base_rent_raw = get_val('base_rent')
        status = str(get_val('status') or 'vacant').strip().lower()

        row_error = None  # เก็บ error message ของแถวนี้

        # Validate ข้อมูลที่จำเป็น
        if not building_name:
            row_error = str(_('building_name is required'))
        elif not floor_number:
            row_error = str(_('floor_number is required'))
        elif not room_number:
            row_error = str(_('room_number is required'))
        else:
            # Validate floor_number เป็นตัวเลข
            try:
                floor_number = int(floor_number)
                if floor_number < 1:
                    raise ValueError
            except (ValueError, TypeError):
                row_error = str(_('floor_number must be a positive integer'))

        if row_error is None:
            # Validate base_rent
            try:
                base_rent = float(base_rent_raw) if base_rent_raw is not None else 0.0
                if base_rent < 0:
                    raise ValueError
            except (ValueError, TypeError):
                row_error = str(_('base_rent must be a non-negative number'))

        if row_error is None:
            # Validate status
            if status not in VALID_ROOM_STATUSES:
                row_error = str(
                    _('status "%(status)s" is invalid. Valid values: %(valid)s') % {
                        'status': status,
                        'valid': ', '.join(sorted(VALID_ROOM_STATUSES)),
                    }
                )

        if row_error is None:
            # ตรวจ duplicate ภายในไฟล์
            key = (building_name.lower(), floor_number, room_number.lower())
            if key in seen_keys:
                row_error = str(
                    _('duplicate entry (%(building)s, Floor %(floor)s, Room %(room)s) in file') % {
                        'building': building_name,
                        'floor': floor_number,
                        'room': room_number,
                    }
                )
            else:
                seen_keys.add(key)

        if row_error is None:
            # ตรวจ duplicate ใน DB (ห้ามซ้ำใน dormitory เดียวกัน)
            exists = Room.unscoped_objects.filter(
                floor__building__dormitory=dormitory,
                floor__building__name=building_name,
                floor__number=floor_number,
                number=room_number,
            ).exists()
            if exists:
                row_error = str(
                    _('room already exists (%(building)s, Floor %(floor)s, Room %(room)s)') % {
                        'building': building_name,
                        'floor': floor_number,
                        'room': room_number,
                    }
                )

        if row_error:
            error_rows.append({'row_num': row_num, 'error': row_error})
        else:
            valid_rows.append({
                'row_num': row_num,
                'building_name': building_name,
                'floor_number': floor_number,
                'room_number': room_number,
                'room_type': room_type,
                'base_rent': base_rent,
                'status': status,
            })

    return valid_rows, error_rows


def _parse_tenant_excel(workbook, dormitory):
    """
    Parse ไฟล์ Excel สำหรับ import ผู้เช่า
    คืนค่า (valid_rows: list[dict], error_rows: list[dict])
    - valid_rows: แถวที่ผ่าน validation ทั้งหมด พร้อม import
    - error_rows: แถวที่มี error พร้อม error message รายแถว
    ทั้งสองรายการ return พร้อมกันเสมอ (partial validation)
    """
    import datetime
    from apps.rooms.models import Room

    ws = workbook.active
    headers = [cell.value for cell in ws[1]]

    # ตรวจสอบว่า header ครบถ้วน — ถ้าไม่ครบให้ abort ทันที
    missing_cols = [col for col in TENANT_IMPORT_COLUMNS if col not in headers]
    if missing_cols:
        fatal_error = _('Missing columns: %(cols)s') % {'cols': ', '.join(missing_cols)}
        return [], [{'row_num': 0, 'error': fatal_error}]

    col_idx = {col: headers.index(col) for col in TENANT_IMPORT_COLUMNS}
    valid_rows = []
    error_rows = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # ข้าม row ว่างทั้งหมด
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        def get_val(col):
            idx = col_idx[col]
            return row[idx] if idx < len(row) else None

        room_number = str(get_val('room_number') or '').strip()
        building_name = str(get_val('building_name') or '').strip()
        first_name = str(get_val('first_name') or '').strip()
        last_name = str(get_val('last_name') or '').strip()
        phone = str(get_val('phone') or '').strip()
        email = str(get_val('email') or '').strip()
        line_id = str(get_val('line_id') or '').strip()
        start_date_raw = get_val('start_date')

        row_error = None  # เก็บ error message ของแถวนี้

        # Validate ข้อมูลที่จำเป็น
        if not room_number:
            row_error = str(_('room_number is required'))
        elif not building_name:
            row_error = str(_('building_name is required'))
        elif not first_name:
            row_error = str(_('first_name is required'))
        elif not last_name:
            row_error = str(_('last_name is required'))

        # Validate start_date — รับได้ทั้ง string YYYY-MM-DD และ date object จาก Excel
        start_date = None
        if row_error is None:
            if isinstance(start_date_raw, (datetime.date, datetime.datetime)):
                start_date = (
                    start_date_raw.date()
                    if isinstance(start_date_raw, datetime.datetime)
                    else start_date_raw
                )
            elif start_date_raw:
                try:
                    start_date = datetime.date.fromisoformat(str(start_date_raw).strip())
                except ValueError:
                    row_error = str(
                        _('start_date "%(val)s" is invalid. Use YYYY-MM-DD format') % {
                            'val': start_date_raw
                        }
                    )
            else:
                row_error = str(_('start_date is required'))

        # ตรวจสอบว่า room มีอยู่ใน dormitory นี้จริง
        room = None
        if row_error is None:
            try:
                room = Room.unscoped_objects.get(
                    floor__building__dormitory=dormitory,
                    floor__building__name=building_name,
                    number=room_number,
                )
            except Room.DoesNotExist:
                row_error = str(
                    _('room "%(room)s" in building "%(building)s" not found in this dormitory') % {
                        'room': room_number, 'building': building_name
                    }
                )
            except Room.MultipleObjectsReturned:
                row_error = str(
                    _('multiple rooms found for "%(room)s" in building "%(building)s"') % {
                        'room': room_number, 'building': building_name
                    }
                )

        # Validate email ถ้ามีการกรอกมา — ตรวจ duplicate ใน DB
        if row_error is None and email:
            from apps.core.models import CustomUser
            if CustomUser.objects.filter(email=email).exists():
                row_error = str(
                    _('email "%(email)s" is already in use') % {'email': email}
                )

        if row_error:
            error_rows.append({'row_num': row_num, 'error': row_error})
        else:
            valid_rows.append({
                'row_num': row_num,
                'room_number': room_number,
                'building_name': building_name,
                'room_id': str(room.pk),
                'first_name': first_name,
                'last_name': last_name,
                'phone': phone,
                'email': email,
                'line_id': line_id,
                'start_date': start_date.isoformat(),
            })

    return valid_rows, error_rows


class ImportRoomsView(OwnerRequiredMixin, View):
    """
    Import ห้องพักจากไฟล์ Excel (.xlsx)
    GET  — แสดง upload form + ปุ่ม download template
    POST action=upload   — parse + validate + preview (เก็บใน session)
    POST action=confirm  — อ่านจาก session → create rooms (atomic)
    POST action=download — ส่ง template Excel กลับทันที
    """

    template_name = 'core/import_rooms.html'

    def get(self, request):
        return render(request, self.template_name, {})

    def post(self, request):
        action = request.POST.get('action', 'upload')

        if action == 'download':
            return self._download_template()
        elif action == 'confirm':
            return self._confirm_import(request)
        else:
            return self._handle_upload(request)

    def _download_template(self):
        """สร้าง Excel template แล้วส่งกลับโดยตรง (ไม่บันทึกไฟล์ไว้)."""
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Rooms'
        ws.append(ROOM_IMPORT_COLUMNS)
        # ใส่ row ตัวอย่าง
        ws.append(['Building A', 1, '101', 'Standard', 5000, 'vacant'])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="rooms_import_template.xlsx"'
        return response

    def _handle_upload(self, request):
        """
        Parse และ validate ไฟล์ Excel แล้วแสดง preview
        แสดงทั้ง valid rows และ error rows — user เลือกเองว่าจะ import เฉพาะ valid หรือ cancel
        """
        import openpyxl

        uploaded_file = request.FILES.get('excel_file')
        if not uploaded_file:
            messages.error(request, _('Please select an Excel file to upload.'))
            return render(request, self.template_name, {})

        if not uploaded_file.name.endswith('.xlsx'):
            messages.error(request, _('Only .xlsx files are supported.'))
            return render(request, self.template_name, {})

        dormitory = request.active_dormitory
        if not dormitory:
            messages.error(request, _('No active dormitory found. Please set up your dormitory first.'))
            return render(request, self.template_name, {})

        try:
            wb = openpyxl.load_workbook(uploaded_file)
        except Exception:
            messages.error(request, _('Could not read the Excel file. Please use a valid .xlsx file.'))
            return render(request, self.template_name, {})

        valid_rows, error_rows = _parse_room_excel(wb, dormitory)

        # ถ้า header ไม่ครบ (fatal error) — แสดง error ทันที ไม่ show preview
        if not valid_rows and error_rows and error_rows[0].get('row_num') == 0:
            return render(request, self.template_name, {
                'fatal_error': error_rows[0]['error'],
            })

        if not valid_rows and not error_rows:
            messages.warning(request, _('No data rows found in the file.'))
            return render(request, self.template_name, {})

        # เก็บเฉพาะ valid rows ใน session สำหรับ confirm step
        # บันทึก dormitory_id ด้วยเพื่อป้องกัน cross-tab dormitory leak
        request.session['import_rooms_preview'] = {
            'dormitory_id': str(dormitory.pk),
            'rows': valid_rows,
        }

        # แสดง preview 10 แถวแรกของ valid rows เท่านั้น
        return render(request, self.template_name, {
            'preview_rows': valid_rows[:10],
            'preview_count': len(valid_rows),
            'error_rows': error_rows,
            'error_count': len(error_rows),
            'has_errors': bool(error_rows),
            'has_valid': bool(valid_rows),
        })

    def _confirm_import(self, request):
        """สร้าง rooms จาก preview data ใน session (atomic transaction)."""
        from apps.rooms.models import Building, Floor, Room

        payload = request.session.get('import_rooms_preview')
        if not payload:
            messages.error(request, _('No import data found. Please upload a file first.'))
            return redirect('core:import_rooms')

        # ป้องกัน cross-tab dormitory leak:
        # ตรวจว่า dormitory ที่ active อยู่ตรงกับ dormitory ที่ upload ไว้ใน session
        dormitory = request.active_dormitory
        if not dormitory:
            messages.error(request, _('No active dormitory found.'))
            return redirect('core:import_rooms')

        if str(dormitory.pk) != payload.get('dormitory_id'):
            messages.error(
                request,
                _('Active dormitory has changed since upload. Please upload the file again.')
            )
            del request.session['import_rooms_preview']
            return redirect('core:import_rooms')

        rows = payload['rows']

        try:
            with transaction.atomic():
                created_count = 0
                for row in rows:
                    building, _created = Building.unscoped_objects.get_or_create(
                        dormitory=dormitory,
                        name=row['building_name'],
                    )
                    floor, _created = Floor.unscoped_objects.get_or_create(
                        building=building,
                        number=row['floor_number'],
                        defaults={'dormitory': dormitory},
                    )
                    Room.unscoped_objects.create(
                        floor=floor,
                        number=row['room_number'],
                        base_rent=row['base_rent'],
                        status=row['status'],
                        dormitory=dormitory,
                    )
                    created_count += 1

            # ล้าง session หลัง import สำเร็จ
            del request.session['import_rooms_preview']
            messages.success(
                request,
                _('Successfully imported %(count)s room(s).') % {'count': created_count}
            )
            return redirect('rooms:list')

        except Exception as exc:
            messages.error(
                request,
                _('Import failed: %(error)s') % {'error': str(exc)}
            )
            return redirect('core:import_rooms')


class ImportTenantsView(OwnerRequiredMixin, View):
    """
    Import ผู้เช่าจากไฟล์ Excel (.xlsx)
    GET  — แสดง upload form + ปุ่ม download template
    POST action=upload   — parse + validate + preview (เก็บใน session)
    POST action=confirm  — อ่านจาก session → create TenantProfiles + Leases (atomic)
    POST action=download — ส่ง template Excel กลับทันที
    """

    template_name = 'core/import_tenants.html'

    def get(self, request):
        return render(request, self.template_name, {})

    def post(self, request):
        action = request.POST.get('action', 'upload')

        if action == 'download':
            return self._download_template()
        elif action == 'confirm':
            return self._confirm_import(request)
        else:
            return self._handle_upload(request)

    def _download_template(self):
        """สร้าง Excel template แล้วส่งกลับโดยตรง."""
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Tenants'
        ws.append(TENANT_IMPORT_COLUMNS)
        # ใส่ row ตัวอย่าง
        ws.append(['101', 'Building A', 'สมชาย', 'ใจดี', '0891234567', 'somchai@email.com', '@somchai', '2026-01-01'])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="tenants_import_template.xlsx"'
        return response

    def _handle_upload(self, request):
        """
        Parse และ validate ไฟล์ Excel แล้วแสดง preview
        แสดงทั้ง valid rows และ error rows — user เลือกเองว่าจะ import เฉพาะ valid หรือ cancel
        """
        import openpyxl

        uploaded_file = request.FILES.get('excel_file')
        if not uploaded_file:
            messages.error(request, _('Please select an Excel file to upload.'))
            return render(request, self.template_name, {})

        if not uploaded_file.name.endswith('.xlsx'):
            messages.error(request, _('Only .xlsx files are supported.'))
            return render(request, self.template_name, {})

        dormitory = request.active_dormitory
        if not dormitory:
            messages.error(request, _('No active dormitory found. Please set up your dormitory first.'))
            return render(request, self.template_name, {})

        try:
            wb = openpyxl.load_workbook(uploaded_file)
        except Exception:
            messages.error(request, _('Could not read the Excel file. Please use a valid .xlsx file.'))
            return render(request, self.template_name, {})

        valid_rows, error_rows = _parse_tenant_excel(wb, dormitory)

        # ถ้า header ไม่ครบ (fatal error) — แสดง error ทันที ไม่ show preview
        if not valid_rows and error_rows and error_rows[0].get('row_num') == 0:
            return render(request, self.template_name, {
                'fatal_error': error_rows[0]['error'],
            })

        if not valid_rows and not error_rows:
            messages.warning(request, _('No data rows found in the file.'))
            return render(request, self.template_name, {})

        # เก็บเฉพาะ valid rows ใน session สำหรับ confirm step
        # บันทึก dormitory_id ด้วยเพื่อป้องกัน cross-tab dormitory leak
        request.session['import_tenants_preview'] = {
            'dormitory_id': str(dormitory.pk),
            'rows': valid_rows,
        }

        # แสดง preview 10 แถวแรกของ valid rows เท่านั้น
        return render(request, self.template_name, {
            'preview_rows': valid_rows[:10],
            'preview_count': len(valid_rows),
            'error_rows': error_rows,
            'error_count': len(error_rows),
            'has_errors': bool(error_rows),
            'has_valid': bool(valid_rows),
        })

    def _confirm_import(self, request):
        """
        สร้าง CustomUser + TenantProfile + Lease จาก preview data ใน session
        ใช้ atomic transaction — rollback ทั้งหมดถ้ามี error ใด row หนึ่ง
        """
        import datetime
        from apps.core.models import CustomUser
        from apps.rooms.models import Room
        from apps.tenants.models import TenantProfile, Lease

        payload = request.session.get('import_tenants_preview')
        if not payload:
            messages.error(request, _('No import data found. Please upload a file first.'))
            return redirect('core:import_tenants')

        # ป้องกัน cross-tab dormitory leak:
        # ตรวจว่า dormitory ที่ active อยู่ตรงกับ dormitory ที่ upload ไว้ใน session
        dormitory = request.active_dormitory
        if not dormitory:
            messages.error(request, _('No active dormitory found.'))
            return redirect('core:import_tenants')

        if str(dormitory.pk) != payload.get('dormitory_id'):
            messages.error(
                request,
                _('Active dormitory has changed since upload. Please upload the file again.')
            )
            del request.session['import_tenants_preview']
            return redirect('core:import_tenants')

        rows = payload['rows']

        try:
            with transaction.atomic():
                created_count = 0
                for row in rows:
                    # สร้าง username จาก first_name + last_name + row_num เพื่อป้องกัน conflict
                    base_username = f"{row['first_name'].lower()}{row['last_name'].lower()}"
                    # ตัดอักษรพิเศษออก
                    import re
                    base_username = re.sub(r'[^a-z0-9]', '', base_username) or f"tenant{row['row_num']}"
                    username = base_username
                    counter = 1
                    while CustomUser.objects.filter(username=username).exists():
                        username = f"{base_username}{counter}"
                        counter += 1

                    # สร้าง user account สำหรับผู้เช่า
                    user = CustomUser.objects.create_user(
                        username=username,
                        email=row['email'],
                        first_name=row['first_name'],
                        last_name=row['last_name'],
                        role=CustomUser.Role.TENANT,
                        dormitory=dormitory,
                    )
                    # ตั้ง password ชั่วคราว = username (ผู้เช่าต้องเปลี่ยนเอง)
                    user.set_password(username)
                    user.save(update_fields=['password'])

                    # ดึง room object
                    room = Room.unscoped_objects.get(pk=row['room_id'])

                    # สร้าง TenantProfile
                    profile = TenantProfile.unscoped_objects.create(
                        user=user,
                        room=room,
                        phone=row['phone'],
                        line_id=row['line_id'],
                        dormitory=dormitory,
                    )

                    # สร้าง Lease (active)
                    Lease.unscoped_objects.create(
                        tenant=profile,
                        room=room,
                        start_date=datetime.date.fromisoformat(row['start_date']),
                        status=Lease.Status.ACTIVE,
                        dormitory=dormitory,
                    )

                    # อัปเดตสถานะห้องเป็น occupied
                    room.status = Room.Status.OCCUPIED
                    room.save(update_fields=['status', 'updated_at'])

                    created_count += 1

            # ล้าง session หลัง import สำเร็จ
            del request.session['import_tenants_preview']
            messages.success(
                request,
                _('Successfully imported %(count)s tenant(s).') % {'count': created_count}
            )
            return redirect('tenants:list')

        except Exception as exc:
            messages.error(
                request,
                _('Import failed: %(error)s') % {'error': str(exc)}
            )
            return redirect('core:import_tenants')


class _WizardForm:
    """Simple wrapper to pass POST data back to template."""

    def __init__(self, step, data=None):
        self.step = step
        self._data = data or {}

    def __getattr__(self, name):
        class _Field:
            def __init__(self, val):
                self._val = val
                self.errors = []

            def value(self):
                return self._val
        return _Field(self._data.get(name, ''))
